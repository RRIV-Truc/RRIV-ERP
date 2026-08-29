"""Parse / publish bảng kế hoạch triển khai TBKL."""
from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional

from modules.meetings.rbac import UserContext
from modules.tbkl import service as svc

_DIRECTIVE_CODE = re.compile(r'^H(\d+)-(\d{2})$', re.I)
_TASK_CODE = re.compile(r'^H(\d+)-(\d{2})-(\d{2})$', re.I)

_PLAN_HEADERS = [
    'Loại', 'Mã', 'Kết luận / Đầu việc', 'SP (sản phẩm)',
    'Phòng chủ trì', 'Đơn vị TH', 'Hạn (YYYY-MM-DD)',
]


def _norm_header(h: str) -> str:
    return re.sub(r'\s+', ' ', (h or '').strip().lower())


def _cell_str(val: Any) -> str:
    if val is None:
        return ''
    if hasattr(val, 'isoformat'):
        try:
            return val.isoformat()[:10]
        except Exception:
            pass
    return str(val).strip()


def _row_kind(raw: dict) -> str:
    kind = _norm_header(raw.get('loại') or raw.get('loai') or raw.get('cap') or '')
    if kind in ('lớn', 'lon', 'directive', 'kết luận', 'ket luan', 'mục lớn'):
        return 'directive'
    if kind in ('chi tiết', 'chi tiet', 'task', 'đầu việc', 'dau viec', 'con'):
        return 'task'
    code = _cell_str(raw.get('mã') or raw.get('ma') or raw.get('code'))
    if _TASK_CODE.match(code):
        return 'task'
    if _DIRECTIVE_CODE.match(code):
        return 'directive'
    return 'task'


def _map_row(cells: list[str], headers: list[str]) -> dict:
    out: dict[str, str] = {}
    for i, h in enumerate(headers):
        key = _norm_header(h)
        val = _cell_str(cells[i]) if i < len(cells) else ''
        out[key] = val
    return out


def _field(row: dict, *keys: str) -> str:
    for k in keys:
        for rk, rv in row.items():
            if k in rk:
                return _cell_str(rv)
    return ''


def _rows_to_plan(rows: list[dict]) -> dict:
    directives: list[dict] = []
    current: Optional[dict] = None

    for row in rows:
        if not any(_cell_str(v) for v in row.values()):
            continue
        kind = _row_kind(row)
        title = _field(row, 'kết luận', 'ket luan', 'đầu việc', 'dau viec', 'nội dung')
        if not title:
            continue

        payload = {
            'title': title,
            'content': title if kind == 'directive' else '',
            'deliverable': _field(row, 'sp', 'sản phẩm', 'san pham'),
            'lead_department_name': _field(row, 'phòng chủ trì', 'phong chu tri'),
            'owner_unit_name': _field(row, 'đơn vị', 'don vi'),
            'deadline': _field(row, 'hạn', 'han') or None,
            'supervisor_name': _field(row, 'giám sát', 'giam sat'),
        }

        if kind == 'directive':
            current = {
                'title': payload['title'],
                'content': payload['content'] or payload['title'],
                'lead_department_name': payload['lead_department_name'] or None,
                'supervisor_name': payload['supervisor_name'] or None,
                'deadline': payload['deadline'],
                'tasks': [],
            }
            directives.append(current)
        else:
            task = {
                'title': payload['title'],
                'deliverable': payload['deliverable'] or None,
                'owner_unit_name': payload['owner_unit_name'] or None,
                'deadline': payload['deadline'],
            }
            if current is None:
                current = {
                    'title': 'Kết luận chung',
                    'content': 'Kết luận chung',
                    'tasks': [],
                }
                directives.append(current)
            current['tasks'].append(task)

    return {'directives': directives}


def parse_csv(data: bytes) -> dict:
    text = data.decode('utf-8-sig', errors='replace')
    reader = csv.reader(io.StringIO(text))
    rows_raw = list(reader)
    if not rows_raw:
        return {'directives': []}
    headers = rows_raw[0]
    body = [_map_row(r, headers) for r in rows_raw[1:]]
    return _rows_to_plan(body)


def parse_xlsx(data: bytes) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('Server thiếu openpyxl — liên hệ quản trị') from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    grid = list(ws.iter_rows(values_only=True))
    if not grid:
        return {'directives': []}
    headers = [_cell_str(c) for c in grid[0]]
    body = []
    for line in grid[1:]:
        cells = [_cell_str(c) for c in line]
        if not any(cells):
            continue
        body.append(_map_row(cells, headers))
    return _rows_to_plan(body)


def parse_workbook(data: bytes, filename: str) -> dict:
    ext = (filename or '').lower().rsplit('.', 1)[-1] if '.' in (filename or '') else ''
    if ext == 'csv':
        return parse_csv(data)
    if ext in ('xlsx', 'xls'):
        if ext == 'xls':
            raise ValueError('Vui lòng lưu file .xls thành .xlsx trước khi tải lên')
        return parse_xlsx(data)
    raise ValueError('Định dạng bảng kế hoạch không hỗ trợ — dùng .xlsx hoặc .csv')


def generate_template_xlsx(meeting_seq: int = 1) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ke hoach TBKL'
    ws.append(_PLAN_HEADERS)
    header_fill = PatternFill('solid', fgColor='D1FAE5')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    samples = [
        ('Lớn', f'H{meeting_seq}-01', 'Thực hiện kế hoạch sản lượng mủ cao su 744 tấn năm 2026', '', 'Phòng KHCN', '', '2026-12-31'),
        ('Chi tiết', f'H{meeting_seq}-01-01', 'Xây dựng phương án điều hành sản lượng', 'Phương án điều hành', 'Phòng KHCN', 'TT Kỹ thuật', '2026-08-31'),
        ('Chi tiết', f'H{meeting_seq}-01-02', 'Thiết lập Dashboard điều hành sản lượng mủ', 'Dashboard cập nhật tuần', 'Phòng KHCN', 'TT Tin học', '2026-09-15'),
        ('Lớn', f'H{meeting_seq}-02', 'Tổ chức và Quy chế hoạt động…', '', 'Phòng TCHC', '', '2026-10-31'),
        ('Chi tiết', f'H{meeting_seq}-02-01', 'Ban hành quy chế phối hợp', 'Quy chế', 'Phòng TCHC', 'Phòng NV', '2026-09-30'),
    ]
    for row in samples:
        ws.append(list(row))

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 48
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def publish_plan(
    supabase,
    ctx: UserContext,
    cycle_id: str,
    plan: dict,
    *,
    replace: bool = False,
) -> dict:
    cycle = svc.get_cycle(supabase, cycle_id)
    if not cycle:
        raise LookupError('Không tìm thấy cuộc họp')
    if cycle.get('status') == 'locked':
        raise ValueError('Cuộc họp đã chốt — không cập nhật kế hoạch')

    existing = supabase.table('tbkl_directives').select('id').eq('cycle_id', cycle_id).execute()
    if existing.data and not replace:
        raise ValueError(
            'Cuộc họp đã có kết luận — tick «Ghi đè» hoặc xóa kết luận cũ trước khi nạp bảng mới'
        )

    if replace and existing.data:
        supabase.table('tbkl_directives').delete().eq('cycle_id', cycle_id).execute()

    directives_in = plan.get('directives') or []
    if not directives_in:
        raise ValueError('Bảng kế hoạch trống — thêm ít nhất một mục kết luận lớn')

    dir_count = 0
    task_count = 0
    for d in directives_in:
        d_payload = {
            'title': (d.get('title') or '').strip() or 'Kết luận',
            'content': (d.get('content') or d.get('title') or '').strip(),
            'lead_department_id': d.get('lead_department_id') or None,
            'lead_department_name': (d.get('lead_department_name') or '').strip() or None,
            'supervisor_name': (d.get('supervisor_name') or '').strip() or None,
            'priority': d.get('priority') or 'normal',
            'deadline': d.get('deadline') or None,
        }
        directive = svc.create_directive(supabase, ctx, cycle_id, d_payload)
        dir_count += 1
        for t in d.get('tasks') or []:
            t_payload = {
                'title': (t.get('title') or '').strip() or 'Đầu việc',
                'deliverable': (t.get('deliverable') or '').strip() or None,
                'owner_unit_id': t.get('owner_unit_id') or None,
                'owner_unit_name': (t.get('owner_unit_name') or '').strip() or None,
                'coordinator_units': (t.get('coordinator_units') or '').strip() or None,
                'assignee_name': (t.get('assignee_name') or '').strip() or None,
                'deadline': t.get('deadline') or None,
                'priority': t.get('priority') or 'normal',
            }
            svc.create_task(supabase, ctx, directive['id'], t_payload)
            task_count += 1

    return {'directive_count': dir_count, 'task_count': task_count}
