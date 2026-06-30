#!/usr/bin/env python3
"""Import danh sách nhân sự Viện từ Excel (BHXH C45) → Supabase.

Mặc định: g:\\My Drive\\Chuyen doi so PTN\\RRIV-ERP\\danh sách nhan su Vien 24.6.xlsx

Chạy: python scripts/import_vien_personnel.py [--dry-run]
"""
from __future__ import annotations

import argparse
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from dotenv import load_dotenv

load_dotenv(ROOT / ".env")

DEFAULT_XLSX = Path(
    r"g:\My Drive\Chuyen doi so PTN\RRIV-ERP\danh sách nhan su Vien 24.6.xlsx"
)

SKIP_B_LABELS = frozenset({
    "họ và tên", "số tt", "nữ",
    "danh sách lao động nộp bhxh, bhyt, bhtn",
})

# Excel tên PB → id chuẩn (sau consolidate_vien_org)
CANONICAL_DEPT_MAP = {
    "ban lãnh đạo": "dl-1",
    "ban lãnh đạo viện": "dl-1",
    "phòng quản trị nhân sự - hành chính": "dl-6",
    "phòng tài chính - kế toán": "dl-6",
    "phòng quản trị tài chính kế toán": "dl-6",
    "phòng kế hoạch và khoa học - công nghệ": "dl-5",
    "trung tâm qlclcs thiên nhiên": "dl-2",
    "trung tâm nghiên cứu phát triển sản phẩm mới": "dl-2",
    "trung tâm nghiên cứu phát triển giống cao su": "dl-3",
    "trung tâm nghiên cứu và chuyển giao tiến bộ kỹ thuật": "dl-4",
    "trung tâm nghiên cứu ứng dụng nông nghiệp công nghệ cao": "dl-4",
    "trung tâm nghiên cứu ứng dụng nông nghiệp công nghệ cao và chuyển giao kỹ thuật": "dl-4",
    "trung tâm ncpt cao su tiểu điền": "dl-3",
    "trung tâm nccgkt tây nguyên": "dl-3",
    "trạm tncs phú yên": "dl-3",
}

CANONICAL_DEPT_NAMES = {
    "dl-1": "Ban Lãnh đạo Viện",
    "dl-5": "Phòng Kế hoạch và Khoa học - Công nghệ",
    "dl-6": "Phòng Quản trị tài chính kế toán",
    "dl-2": "Trung tâm nghiên cứu phát triển sản phẩm mới",
    "dl-3": "Trung tâm nghiên cứu phát triển Giống cao su",
    "dl-4": "Trung tâm nghiên cứu ứng dụng nông nghiệp công nghệ cao",
}


def canonical_dept(name: str, order: int) -> dict:
    key = name.strip().lower()
    dept_id = CANONICAL_DEPT_MAP.get(key, f"vien-{order:02d}")
    canon_name = CANONICAL_DEPT_NAMES.get(dept_id, name.strip())
    return {"name": canon_name, "order": order, "id": dept_id}


def is_stt_cell(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    if not s:
        return False
    try:
        int(float(s))
        return True
    except ValueError:
        return False


def is_department_name(name: str) -> bool:
    s = name.strip()
    if len(s) < 4:
        return False
    sl = s.lower()
    if sl in SKIP_B_LABELS:
        return False
    if "tập đoàn" in sl or "viện nghiên cứu cao su" in sl:
        return False
    if sl.startswith("ban ") or sl.startswith("ban\t") or sl.startswith("ban l"):
        return True
    if sl.startswith("phòng"):
        return True
    if "trung tâm" in sl or sl.startswith("trung t"):
        return True
    if sl.startswith("trạm"):
        return True
    return False


def is_department_row(a, b, c, d, e, f) -> bool:
    if not b or is_stt_cell(a):
        return False
    s = str(b).strip()
    if not is_department_name(s):
        return False
    # Dòng tiêu đề PB: thường không có ngày sinh / CCCD / chức vụ trên cùng hàng
    if c or e or f:
        return False
    return True


def parse_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    if " " in s:
        return s.split(" ")[0][:10]
    return s[:10] if len(s) >= 10 else None


def parse_excel(path: Path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    depts: list[dict] = []
    people: list[dict] = []
    current_dept: str | None = None

    for row in ws.iter_rows(values_only=True):
        cols = list(row) + [None] * 6
        a, b, c, d, e, f = cols[:6]

        if is_department_row(a, b, c, d, e, f):
            s = str(b).strip()
            current_dept = s
            order = len(depts) + 1
            depts.append(canonical_dept(s, order))
            continue

        if is_stt_cell(a) and b:
            stt = int(float(str(a).strip()))
            dept_id = depts[-1]["id"] if depts else None
            people.append({
                "stt": stt,
                "name": str(b).strip(),
                "dob": parse_date(c),
                "gender": "Female" if (d and str(d).strip()) else "Male",
                "position": str(e).strip() if e else "",
                "cccd": re.sub(r"\D", "", str(f)) if f else "",
                "department_name": depts[-1]["name"] if depts else current_dept,
                "department_id": dept_id,
            })
    wb.close()
    return depts, people


def dept_type(name: str) -> str:
    n = name.lower()
    if n.startswith("ban"):
        return "Ban Lãnh Đạo"
    if n.startswith("phòng"):
        return "Phòng Nghiệp Vụ"
    return "Trung Tâm"


def skip_institute_import(position: str) -> bool:
    """Chỉ bỏ KH — CN vẫn thuộc danh sách Viện."""
    return "khoán hộ" in (position or "").lower()


DEPT_CODE_SUFFIX = {
    "dl-1": "01",
    "dl-5": "03",
    "dl-6": "02",
    "dl-2": "05",
    "dl-3": "06",
    "dl-4": "07",
    "vien-01": "01",
    "vien-03": "03",
    "vien-05": "05",
    "vien-06": "06",
    "vien-07": "07",
}


def make_code(dept_id: str, stt: int) -> str:
    suffix = DEPT_CODE_SUFFIX.get(dept_id) or dept_id.split("-")[1]
    return f"VIEN-{suffix}-{stt:03d}"


def find_existing_employee(db, person: dict, code: str):
    if person["cccd"]:
        res = (
            db.table("employee")
            .select("id, employee_code")
            .eq("national_id", person["cccd"])
            .limit(1)
            .execute()
        )
        if res.data:
            return res.data[0]

    res = (
        db.table("employee")
        .select("id, employee_code")
        .eq("employee_code", code)
        .limit(1)
        .execute()
    )
    if res.data:
        return res.data[0]

    if person["department_name"]:
        res = (
            db.table("employee")
            .select("id, employee_code")
            .eq("full_name", person["name"])
            .eq("department_name", person["department_name"])
            .limit(1)
            .execute()
        )
        if res.data:
            return res.data[0]
    return None


def code_taken_by_other(db, code: str, employee_id: str | None = None) -> bool:
    query = db.table("employee").select("id").eq("employee_code", code)
    if employee_id:
        query = query.neq("id", employee_id)
    res = query.limit(1).execute()
    return bool(res.data)


def next_free_code(db, base_code: str, employee_id: str | None = None) -> str:
    if not code_taken_by_other(db, base_code, employee_id):
        return base_code
    m = re.match(r"^(.+)-(\d+)$", base_code)
    prefix = m.group(1) if m else base_code
    for n in range(2, 1000):
        candidate = f"{prefix}-{n:02d}"
        if not code_taken_by_other(db, candidate, employee_id):
            return candidate
    return f"{base_code}-{uuid.uuid4().hex[:6].upper()}"


def import_to_supabase(depts: list[dict], people: list[dict], dry_run: bool = False) -> None:
    import os
    from supabase import create_client

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_KEY")
    if not url or not key:
        print("Thiếu SUPABASE_URL / SUPABASE_KEY trong .env")
        sys.exit(1)

    if dry_run:
        print(f"[dry-run] {len(depts)} phòng ban, {len(people)} nhân sự")
        for d in depts:
            print(f"  {d['order']:2d}. {d['name']}")
        return

    db = create_client(url, key)

    for d in depts:
        row = {
            "id": d["id"],
            "name": d["name"],
            "ten": d["name"],
            "ten_phong_ban": d["name"],
            "dept_type": dept_type(d["name"]),
            "active": True,
            "metadata": {
                "order": d["order"],
                "org_type": "vien",
                "source": "excel-vien-24.6",
            },
        }
        db.table("category_departments").upsert(row).execute()

    created = updated = skipped = errors = 0
    for p in people:
        if not p["department_id"]:
            skipped += 1
            continue
        if skip_institute_import(p.get("position", "")):
            skipped += 1
            continue

        code = make_code(p["department_id"], p["stt"])
        national_id = p["cccd"] or None
        meta = {
            "listStt": p["stt"],
            "orderByDept": {p["department_id"]: p["stt"]},
            "source": "excel-vien-24.6",
        }
        patch = {
            "full_name": p["name"],
            "gender": p["gender"],
            "date_of_birth": p["dob"],
            "department_id": p["department_id"],
            "department_name": p["department_name"],
            "position_name": p["position"],
            "employment_status": "active",
            "disabled": False,
            "metadata": meta,
        }
        if national_id:
            patch["national_id"] = national_id

        try:
            existing = find_existing_employee(db, p, code)
            if existing:
                eid = existing["id"]
                # Giữ mã cũ nếu mã Excel đã thuộc người khác
                if code_taken_by_other(db, code, eid):
                    if existing.get("employee_code"):
                        pass  # không đổi employee_code
                    else:
                        patch["employee_code"] = next_free_code(db, code, eid)
                else:
                    patch["employee_code"] = code
                db.table("employee").update(patch).eq("id", eid).execute()
                updated += 1
            else:
                insert_code = next_free_code(db, code)
                row = {
                    **patch,
                    "employee_code": insert_code,
                    "national_id": national_id or f"MIG-{insert_code}",
                    "id": str(uuid.uuid4()),
                }
                db.table("employee").insert(row).execute()
                created += 1
        except Exception as exc:
            errors += 1
            print(f"  Lỗi [{p['name']} / {code}]: {exc}")

    print(
        f"Phòng ban: {len(depts)} | Tạo mới: {created} | "
        f"Cập nhật: {updated} | Bỏ qua: {skipped} | Lỗi: {errors}"
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"Không tìm thấy file: {args.xlsx}")
        sys.exit(1)

    depts, people = parse_excel(args.xlsx)
    print(f"Đọc được {len(depts)} phòng ban / trung tâm, {len(people)} nhân sự")
    import_to_supabase(depts, people, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
